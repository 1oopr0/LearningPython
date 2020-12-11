from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "SeungMinSheet"

ws["A1"] = 1 # A1 셀에 1 이라는 값을 입력
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])         # 셀 정보
print(ws["A1"].value)   # 셀 '값' (없으면 None을 출력함)

ws.cell(row=1, column=1) # A1 셀과 같음

wb.save("sample.xlsx")

